#from undetected_chromedriver   import Chrome, ChromeOptions
import pandas as pd
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

import os
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import range as xlrange
from datetime import date, timedelta

from datetime import datetime
from dateutil.relativedelta import relativedelta
import calendar

chrome_options = webdriver.ChromeOptions()
#chrome_options.add_argument("--headless")
#chrome_options.add_argument("--disable-gpu")

#driver.execute_script("document.body.style.zoom='80%'")
def find_element_click(path,browser):
    while True:
        try:
            browser.find_element(By.XPATH,path).click()
            break
        except Exception as e:
            print(e)


    return browser

def find_element_key(path,sendkey,browser):
    while True:
        try:
            browser.find_element(By.XPATH,path).send_keys(sendkey)
            break
        except Exception as e:
            print(e)

    return browser


def find_element_key_enter(path,sendkey,browser):
    while True:
        try:
            b=browser.find_element(By.XPATH,path).send_keys(sendkey).send_keys('Keys.ENTER')
            time.sleep(1)
            b.click()
            break
        except Exception as e:
            print(e)

    return browser
# Función para comprobar si una fecha es lunes
def es_lunes(fecha):
    # Usa la función strftime() para obtener el día de la semana como un número (0: lunes, 1: martes, ..., 6: domingo)
    
    #fecha= datetime.strptime(fecha, '%Y-%m-%d %H:%M:%S ')
    
    dia_semana = fecha.strftime('%w')
    # Comprueba si el día de la semana es 1 (lunes)
    return dia_semana == '1'


def obtener_primer_ultimo_dia_mes(anio, mes):
    primer_dia = 1
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    return primer_dia, ultimo_dia

# Función para dar formato a la fecha en 'a b d Y'
def dar_formato_fecha(anio, mes, dia):
    fecha = datetime(anio, mes, dia)
    return fecha.strftime('%a %b %d %Y')
  
def configuration_youbora():
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')    
    #chrome_options.add_argument("--force-device-scale-factor=0.9")

    # si 31(hoy)-7 = 24 no
    #    7(hoy) -7 = 31 si
    #    14(hoy) -7 =7  no
    #    27 - 7 =20 no  
    #    3- 7 = 27 si  
    #    1 -7 =24 si
    #    4 -7 =28 si  
    #    2-7 =25 si
    
    time.sleep(1)
    path_update_mes =['https://suite.npaw.com/v/ob-peru/analytics/d/133886','https://suite.npaw.com/v/ob-peru/analytics/d/133879','https://suite.npaw.com/v/ob-peru/analytics/d/133880','https://suite.npaw.com/v/ob-peru/analytics/d/133716']
    for name_path in path_update_mes:
        chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
        #Crear el objeto Service con la ruta del controlador
        service = Service(chrome_driver_path)
        driver = webdriver.Chrome(service=service,options=options)
        print(name_path)
        time.sleep(1)
        driver.get(str(name_path))
        time.sleep(1)
        driver.maximize_window()
        find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
        find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)        
        find_element_click('//*[@id="youbora__login_submit"]',driver)

        FECHA = datetime.now()#-timedelta(days=7)
        dia = FECHA.day      
       
        print(name_path)
        print(dia)

        if dia<=7:
            time.sleep(2)
            FECHA = datetime.now()-timedelta(days=7)
            print(FECHA)

            anio_actual = FECHA.year
            mes_actual = FECHA.month
            primer_dia, ultimo_dia = obtener_primer_ultimo_dia_mes(anio_actual, mes_actual)

            FECHA_INIT = dar_formato_fecha(anio_actual, mes_actual, primer_dia)
            FECHA_FIN = dar_formato_fecha(anio_actual, mes_actual, ultimo_dia)
            print(FECHA_INIT)
            print(FECHA_FIN)

            # dia  = str(FECHA_INIT.strptime('%a %b %d %Y'))
            # dia_f  = str(FECHA_FIN.strptime('%a %b %d %Y'))
            find_element_click("/html/body/div[2]/main/header/div[2]/div[1]/div[1]/div/div[2]/input",driver)

            DIA='//*[@id="date-picker-calendar"]//div[@aria-label="'+str(FECHA_INIT)+'"]' 
            DIA_FIN='//*[@id="date-picker-calendar"]//div[@aria-label="'+str(FECHA_FIN)+'"]' 
            print(DIA)
            print(DIA_FIN)
            time.sleep(2)

            find_element_click(DIA,driver)
            find_element_click(DIA_FIN,driver)

            time.sleep(2)
            #Aplicar
            find_element_click('//*[@id="date-picker-calendar"]/div[3]/div[3]/button[2]',driver)

            find_element_click('/html/body/div[2]/main/header/div[2]/div[5]/div[2]/button',driver)
            time.sleep(3)

        driver.close()

    #Configurar si el dia es lunes para mayor seguridad
    path_update_date= ['https://suite.npaw.com/v/ob-peru/analytics/d/148189','https://suite.npaw.com/v/ob-peru/analytics/d/148200','https://suite.npaw.com/v/ob-peru/analytics/d/148211','https://suite.npaw.com/v/ob-peru/analytics/d/148212']
    for name_path in path_update_date:
        chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
        # Crear el objeto Service con la ruta del controlador
        service = Service(chrome_driver_path)
        driver = webdriver.Chrome(service=service,options=options)
        print(name_path)
        time.sleep(1)
        driver.get(str(name_path))
        time.sleep(1)
        driver.maximize_window()
        find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
        find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)        
        find_element_click('//*[@id="youbora__login_submit"]',driver)
        time.sleep(2)   
        find_element_click("/html/body/div[2]/main/header/div[2]/div[1]/div[1]/div/div[2]/input",driver)        
        FECHA_INIT = datetime.now()-timedelta(days=7)
        FECHA_FIN =  FECHA_INIT + timedelta(days=6)        
        d=FECHA_INIT.strftime('%Y-%m-%d %H:%M:%S')
        print(d)
        dia  = str(FECHA_INIT.strftime('%a %b %d %Y'))
        dia_f  = str(FECHA_FIN.strftime('%a %b %d %Y'))
        print(dia)
        DIA='//*[@id="date-picker-calendar"]//div[@aria-label="'+dia+'"]' 
        DIA_FIN='//*[@id="date-picker-calendar"]//div[@aria-label="'+dia_f+'"]' 
        print(DIA)
        time.sleep(0.5)
        find_element_click(DIA,driver)
        find_element_click(DIA_FIN,driver)
        time.sleep(1)
        #Aply
        find_element_click('//*[@id="date-picker-calendar"]/div[3]/div[3]/button[2]',driver)
        #guardar
        time.sleep(1)
        find_element_click('/html/body/div[2]/main/header/div[2]/div[5]/div[2]/button',driver)
        time.sleep(3)
        driver.close()

  


def configuration_date():
    # Obtener la fecha actual
    fecha_actual = datetime.now()
    print(fecha_actual)

    if es_lunes(fecha_actual):
        print(f"La fecha {fecha_actual} es un lunes.")
        configuration_youbora()
    else:
        print(f"La fecha {fecha_actual} no es un lunes.")

def filtro1_live():

    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    
    prefs = {'download.default_directory' : r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter1'}
    options.add_experimental_option('prefs', prefs)
    #chrome_options.add_argument("--force-device-scale-factor=0.9")

    chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'

   # Crear el objeto Service con la ruta del controlador
    service = Service(chrome_driver_path)


    driver = webdriver.Chrome(service=service,options=options)
    #driver = webdriver.Chrome(executable_path="C:\dina2\dina2\src\chromedriver.exe",chrome_options=options)

    driver.maximize_window()

    driver.get("https://suite.npaw.com/v/ob-peru/analytics/d/148189")
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)
    
    find_element_click('//*[@id="youbora__login_submit"]',driver)
    time.sleep(1)

    #UNIQUE USERS
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[2]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #EFFECTIVE PLAYTIME
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[1]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #PLAYS
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[3]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #BUFFER RATIO
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[8]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)  
    time.sleep(1)

    #In-Stream Error Crash
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[15]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #Avg. Bitrate (Mbps)
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[9]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)  
    time.sleep(1)

    #Happiness Score
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[13]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)  

    time.sleep(3)  
    #driver.quit()
    return driver

def filtro2_live():

    options = webdriver.ChromeOptions()    
    prefs = {'download.default_directory' : r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter2'}
    options.add_experimental_option('prefs', prefs)

    #chrome_driver_path='chromedriver.exe'
    #driver = webdriver.Chrome(executable_path=chrome_driver_path,chrome_options=options)
    
    chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service,options=options)
 
    driver.maximize_window()
    driver.get("https://suite.npaw.com/v/ob-peru/analytics/d/148200")
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)    
    find_element_click('//*[@id="youbora__login_submit"]',driver)
    time.sleep(2)
                       
    #PLAYS
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[3]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(3)
                        
    # #BUFFER RATIO
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[8]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver) 
    time.sleep(3)

    # #In-Stream Error Crash    
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[15]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)

    time.sleep(3)

    #driver.quit()

    return driver

def filtro3_live():

    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    
    prefs = {'download.default_directory' : r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter3'}
    options.add_experimental_option('prefs', prefs)
    #chrome_options.add_argument("--force-device-scale-factor=0.9")

    chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service,options=options)

    driver.maximize_window()

    driver.get("https://suite.npaw.com/v/ob-peru/analytics/d/148211")
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)
    
    find_element_click('//*[@id="youbora__login_submit"]',driver)
    time.sleep(1)
    
    #In-Stream Error Crash    
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[15]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)

    time.sleep(1)


    #PLAYS
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[3]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    
    time.sleep(1)


    #BUFFER RATIO
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[8]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver) 
    time.sleep(3)

    return driver
    #driver.quit()
def filtro4_live():

    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    
    prefs = {'download.default_directory' : r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter4'}
    options.add_experimental_option('prefs', prefs)
    #chrome_options.add_argument("--force-device-scale-factor=0.9")

    chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service,options=options)

    driver.maximize_window()

    driver.get("https://suite.npaw.com/v/ob-peru/analytics/d/148212")
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)
    
    find_element_click('//*[@id="youbora__login_submit"]',driver)

    #UNIQUE USERS
    
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[2]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #EFFECTIVE PLAYTIME
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[1]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #PLAYS
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[3]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #BUFFER RATIO
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[8]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)  
    time.sleep(1)

    #In-Stream Error Crash
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[15]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
    time.sleep(1)

    #Avg. Bitrate (Mbps)
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[9]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)  
    time.sleep(1)

    #Happiness Score
    find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[13]/div/div[1]/div[2]/div[1]/div',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)  

    time.sleep(3)
    #driver.quit()
  
    return driver

def filtro5_live():
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    filters = [
    {'filter':r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter5','path':'https://suite.npaw.com/v/ob-peru/analytics/d/148212'},
    {'filter':r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter6','path':'https://suite.npaw.com/v/ob-peru/analytics/d/148211'},
    {'filter':r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter7','path':'https://suite.npaw.com/v/ob-peru/analytics/d/148200'},
    ]

    for i in filters:
        prefs = {'download.default_directory' : i['filter']}   
        # prefs = {'download.default_directory' : r'C:\Users\wrivera\Documents\RE__Reporte_Performance\filter5'}
        options.add_experimental_option('prefs', prefs)
        # chrome_driver_path='chromedriver.exe'
        # driver = webdriver.Chrome(executable_path=chrome_driver_path,chrome_options=options)
        chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
        service = Service(chrome_driver_path)
        driver = webdriver.Chrome(service=service,options=options)

        driver.maximize_window()
        driver.get(i['path'])

        find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
        find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)
    
        find_element_click('//*[@id="youbora__login_submit"]',driver)
        time.sleep(5)
        
        FECHA_INIT = datetime.now()#-timedelta(days=2)
        FECHA_INIT =  FECHA_INIT - timedelta(days=7)
        #print(FECHA_INIT)
        #FECHA_INIT='2023-07-10 00:00:00'


        if es_lunes(FECHA_INIT):
            print(f"La fecha {FECHA_INIT} es un lunes.")
            #configuration_youbora()
            
            dias = 7    
            # Rellenar las fechas en la columna A  
            for i in range(dias):   
                find_element_click("/html/body/div[2]/main/header/div[2]/div[1]/div[1]/div/div[2]/input",driver)

                #d= FECHA_INIT
                #d = datetime.strptime(FECHA_INIT, '%Y-%m-%d %H:%M:%S')
                #dia  = str(d.strftime('%a %b %d %Y'))
                #d = datetime.strftime(FECHA_INIT, '%Y-%m-%d')

                d=FECHA_INIT.strftime('%Y-%m-%d %H:%M:%S')
                print('ddddddd '+d)
                dia  = str(FECHA_INIT.strftime('%a %b %d %Y'))

                # dia  = str(FECHA_INIT.day)

                print('diaaaa '+str(dia))

                DIA='//*[@id="date-picker-calendar"]//div[@aria-label="'+dia+'"]' 

                print(DIA)
                time.sleep(0.5)

                find_element_click(DIA,driver)
                find_element_click(DIA,driver)
                time.sleep(0.5)
                #aplicar                                                
                find_element_click('/html/body/div[4]/div[3]/div[3]/button[2]',driver)
                time.sleep(0.5)
                #combo
                find_element_click('/html/body/div[2]/main/header/div[2]/div[2]/div/div/div[2]/div[2]/div',driver)
                time.sleep(0.5)
                

                #find_element_click('/html/body/div[2]/main/header/div[2]/div[2]/div/div/div[2]/div[1]/div[1]/p[1]',driver)

                find_element_key('/html/body/div[2]/main/header/div[2]/div[2]/div/div/div[2]/div[1]/div[2]/input',"Horas",driver)
                                
                time.sleep(0.5)

                
                find_element_click('/html/body/div[2]/main/div[2]/div[1]/div[4]/div/div[1]/div[2]/div[1]/div',driver)
                time.sleep(0.5)
              
                find_element_click('/html/body/div[4]/div[3]/ul/li[5]',driver)
                time.sleep(0.5)

                find_element_click('/html/body/div[4]/div[3]/ul/div/div/div/li[1]',driver)
                time.sleep(0.5)

                FECHA_INIT =FECHA_INIT +  timedelta(days=1)
                print(FECHA_INIT)
                time.sleep(2)

            driver.close()
        else:
            print(f"La fecha {FECHA_INIT} no es un lunes.")

    time.sleep(2)
    #return driver
    driver.quit()
        

def filtro_mensual():
    options = webdriver.ChromeOptions()    
    prefs = {'download.default_directory' : r'C:\Users\wrivera\Documents\RE__Reporte_Performance\files_mensual'}
    options.add_experimental_option('prefs', prefs)

    # chrome_driver_path='chromedriver.exe'
    # driver = webdriver.Chrome(executable_path=chrome_driver_path,chrome_options=options)
    chrome_driver_path=r'C:\Users\wrivera\Documents\Documents\RE__Reporte_Performance\chromedriver.exe'
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service,options=options)

 
    driver.maximize_window()
    driver.get("https://suite.npaw.com/v/ob-peru/analytics/d/133886")
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[1]/div/input',"PeruOps",driver)
    find_element_key('//*[@id="youbora__container"]/div[1]/form/div[2]/div/input',"P3ru0ps",driver)    
    find_element_click('//*[@id="youbora__login_submit"]',driver)
    time.sleep(1)
 

    find_element_click('/html/body/div[2]/main/header/div[2]/div[5]/div[4]/button',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/li',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/div/div/div/li[1]',driver)
    time.sleep(3)

    driver.get('https://suite.npaw.com/v/ob-peru/analytics/d/133879')

    find_element_click('/html/body/div[2]/main/header/div[2]/div[5]/div[4]/button',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/li',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/div/div/div/li[1]',driver)
    time.sleep(3)

    driver.get('https://suite.npaw.com/v/ob-peru/analytics/d/133880')

    find_element_click('/html/body/div[2]/main/header/div[2]/div[5]/div[4]/button',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/li',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/div/div/div/li[1]',driver)
    time.sleep(3)

    driver.get('https://suite.npaw.com/v/ob-peru/analytics/d/133716')
    find_element_click('/html/body/div[2]/main/header/div[2]/div[5]/div[4]/button',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/li',driver)
    find_element_click('/html/body/div[4]/div[3]/ul/div[6]/div/div/div/li[1]',driver)
    time.sleep(3)

from openpyxl.utils import get_column_letter

def div_results(datos_transpuestos):
    print(datos_transpuestos)
    return 1

def read_datos_entrada(archivo_entrada,fila_inicio,columna_inicio):
    # Leer los datos del archivo CSV de entrada a partir de la fila y columna especificadas
    datos_entrada = []
    with open(archivo_entrada, "r") as f:
        lector_csv = csv.reader(f)
       
        for _ in range(fila_inicio - 1):  # Omitir filas antes de la fila de inicio
            next(lector_csv)
        for fila in lector_csv:
            datos_entrada.append(fila[columna_inicio - 1:])  # Omitir columnas antes de la columna de inicio

    return datos_entrada

def read_live_mensual(archivo_entrada,fila_inicio,columna_inicio):
    with open(archivo_entrada, 'r') as f:
        lector_csv = csv.reader(f)
        datos_csv = list(lector_csv)  # Convertir el lector CSV en una lista de listas
    
        for fila in datos_csv:
            print(fila)
            # dato1 = fila[0]  # Acceder al primer dato en la fila
            # dato2 = fila[2]  # Acceder al tercer dato en la fila
            # dato3 = fila[4]  # Acceder al quinto dato en la fila
        return fila
        # Realizar las operaciones necesarias con los datos
        
        # Ejemplo: Imprimir los datos
        #print(f"Dato 1: {dato1}, Dato 2: {dato2}, Dato 3: {dato3}")

def lector_valores(archivo_entrada,fila_inicio,columna_inicio):
    with open(archivo_entrada, 'r') as f:
        lector_csv = csv.reader(f)    
        numeros = []
        # Recorrer las filas del CSV
        for fila in lector_csv:
            # Extraer los números de las columnas
            #print(fila)
            for valor in fila:
                try:
                    numero = float(valor)
                    numeros.append(numero)
                except ValueError:
                    pass
    return numeros

def excluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_excluir,fin_excluir):
    #filas_mostrar = [23, 3,4, 5] 
    
    with open(archivo_entrada, 'r') as f:
        lector_csv = csv.reader(f)    
        numeros = []
        datos_entrada=[]
        # Recorrer las filas del CSV
        for i,fila in enumerate(lector_csv):
            #if i not in filas_mostrar:
            if i < inicio_excluir or i > fin_excluir:
                #print(fila)
                for valor in fila:
                    try:
                        numero = float(valor)
                        print(numero)
                        numeros.append(numero)
                    except ValueError:
                        pass
                #datos_entrada.append(fila)
    return numeros

def incluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_incluir,fin_incluir):
    columna_inicio = 2
    #filas_mostrar = [23, 3,4, 5] 
  
    with open(archivo_entrada, 'r') as f:
        lector_csv = csv.reader(f)    
        numeros = []
        datos_entrada=[]
        # Recorrer las filas del CSV
        for i,fila in enumerate(lector_csv):
            #if i not in filas_mostrar:
            if i >= inicio_incluir and i <= fin_incluir: 
                datos_entrada.append(fila[columna_inicio - 1:])
    return datos_entrada

def rw_transponer(archivo_entrada,archivo_salida,columna_destino_index,fila_destino,folder):
    hoja_nombre = "LIVE_diario"
    fila_inicio = 4  # Fila desde la cual iniciar la transposición o lectura
    columna_inicio = 2  # Columna desde la cual iniciar la transposición o lectura
    #columna_destino = "AWB"  # Columna de destino donde se insertarán los datos transpuestos #AWB = 1276

    datos_entrada = read_datos_entrada(archivo_entrada,fila_inicio,columna_inicio)

    # Transponer los datos
    datos_transpuestos = list(map(list, zip(*datos_entrada)))

    print(columna_destino_index)


    try:
        libro_salida = load_workbook(archivo_salida)
        nombres_hojas = libro_salida.sheetnames 
        hoja_indice = 3
        hoja_salida = libro_salida[nombres_hojas[hoja_indice]] 
    except FileNotFoundError:
        libro_salida = Workbook()
        hoja_salida = libro_salida.active
        hoja_salida.title = hoja_nombre


    for i, fila_transpuesta in enumerate(datos_transpuestos):
        for j, valor in enumerate(fila_transpuesta):
            if i==1 and folder=="filter3":
                celda = hoja_salida.cell(row=fila_destino+8, column=columna_destino_index + j, value=valor)
                celda.value = valor
                celda.data_type = "n"    
            else:
                celda = hoja_salida.cell(row=fila_destino+i, column=columna_destino_index + j, value=valor)
                celda.value = valor
                celda.data_type = "n"

    # # # Guardar el libro de trabajo de salida en formato XLSX
    libro_salida.save(archivo_salida)

def plantilla():
    # Cargar el libro de Excel existente
    carpeta2 = "Entregable"
    archivo2 = "Indicadores de calidad.xlsx"
    archivo_salida = os.path.join(carpeta2, archivo2)

    #archivo_entrada = os.path.join("filter5", "Plays vs Join Time.csv")
    hoja_nombre = "LIVE_diario"
    fila_inicio = 4  # Fila desde la cual iniciar la transposición o lectura
    columna_inicio = 2  # Columna desde la cual iniciar la transposición o lectura
    #columna_destino_index = 1283
    fila=2
    columna_destino_index=verify_column_empty(archivo_salida,hoja_nombre,fila)
    columna_destino_index=columna_destino_index-7
    print(columna_destino_index)

    column_name=get_column_letter(columna_destino_index)
    print(column_name)

    # Cargar el libro de trabajo de salida (si existe) o crear uno nuevo
    try:
        libro_salida = load_workbook(archivo_salida)
        nombres_hojas = libro_salida.sheetnames 
        hoja_indice = 3
        hoja_salida = libro_salida[nombres_hojas[hoja_indice]] 
    except FileNotFoundError:
        libro_salida = Workbook()
        hoja_salida = libro_salida.active
        hoja_salida.title = hoja_nombre

    diccionarios = [
            {"name":"Plays vs Join Time.csv", "fila_destino": 107,"folder":"filter5"},
            {"name":"Plays vs Join Time (1).csv", "fila_destino" : 107,"folder":"filter5"},
            {"name":"Plays vs Join Time (2).csv", "fila_destino" : 107,"folder":"filter5"},
            {"name":"Plays vs Join Time (3).csv", "fila_destino" : 107,"folder":"filter5"},
            {"name":"Plays vs Join Time (4).csv", "fila_destino" : 107,"folder":"filter5"},
            {"name":"Plays vs Join Time (5).csv", "fila_destino" : 107,"folder":"filter5"},
            {"name":"Plays vs Join Time (6).csv", "fila_destino" : 107,"folder":"filter5"},
            {"name":"Plays vs Join Time.csv", "fila_destino": 114,"folder":"filter6"},
            {"name":"Plays vs Join Time (1).csv", "fila_destino" : 114,"folder":"filter6"},
            {"name":"Plays vs Join Time (2).csv", "fila_destino" : 114,"folder":"filter6"},
            {"name":"Plays vs Join Time (3).csv", "fila_destino" : 114,"folder":"filter6"},
            {"name":"Plays vs Join Time (4).csv", "fila_destino" : 114,"folder":"filter6"},
            {"name":"Plays vs Join Time (5).csv", "fila_destino" : 114,"folder":"filter6"},
            {"name":"Plays vs Join Time (6).csv", "fila_destino" : 114,"folder":"filter6"},
            {"name":"Plays vs Join Time.csv", "fila_destino": 121,"folder":"filter7"},
            {"name":"Plays vs Join Time (1).csv", "fila_destino" : 121,"folder":"filter7"},
            {"name":"Plays vs Join Time (2).csv", "fila_destino" : 121,"folder":"filter7"},
            {"name":"Plays vs Join Time (3).csv", "fila_destino" : 121,"folder":"filter7"},
            {"name":"Plays vs Join Time (4).csv", "fila_destino" : 121,"folder":"filter7"},
            {"name":"Plays vs Join Time (5).csv", "fila_destino" : 121,"folder":"filter7"},
            {"name":"Plays vs Join Time (6).csv", "fila_destino" : 121,"folder":"filter7"},
    ]
    k=0
    for diccionario in diccionarios:
        archivo_entrada = os.path.join(diccionario['folder'], diccionario['name'])
        #rw_transponer(archivo_entrada,archivo_salida,columna_destino_index,diccionario['fila_destino'],diccionario['folder'])

        datos_entrada = read_datos_entrada(archivo_entrada,fila_inicio,columna_inicio)
        #print(datos_entrada)
        if diccionario['folder']=="filter5":
            acumulado = sum(int(elemento[0]) for elemento in datos_entrada[5:])
            nueva_entrada = datos_entrada[:5] + [[str(acumulado)]]
            
        elif diccionario['folder']=="filter6":
            #ejemplo = [['7148', '0'], ['36918', '13'], ['8289', '1'], ['3139', '0'], ['1315', '0'], ['573', '0'], ['311', '0'], ['175', '0'], ['187', '0'], ['85', '0'], ['54', '0'], ['124', '0']]
            nueva_entrada = datos_entrada[:5]  # Copiar los primeros cinco elementos de la lista original
            sumatorio=[]
            i=0
            acumulado = 0                 

            for i in range(2):
                acumulado = 0                 
                for valor in datos_entrada[5:]:
                    acumulado += int(valor[i])
                sumatorio.append(str(acumulado))                
            nueva_entrada.append(sumatorio)
        
        elif diccionario['folder']=="filter7":
        #entrada = [['5167', '1615', '69', '0', '0', '0', '5'], ['31963', '2619', '773', '0', '0', '13', '5'], ['6671', '668', '496', '0', '0', '1', '130'], ['2463', '207', '142', '0', '0', '0', '205'], ['948', '119', '84', '0', '0', '0', '108'], ['414', '52', '25', '0', '0', '0', '60'], ['244', '20', '20', '0', '0', '0', '10'], ['126', '17', '17', '0', '0', '0', '7'], ['129', '21', '19', '0', '0', '0', '6'], ['57', '10', '10', '0', '0', '0', '2'], ['39', '7', '7', '0', '0', '0', '0'], ['53', '34', '10', '0', '0', '0', '23']]
            nueva_entrada = datos_entrada[:5]  # Copiar los primeros cinco elementos de la lista original            
            acumulado = 0
            sumatorio=[]                         
            i=0
            for i in range(6):
                acumulado=0
                for valor in datos_entrada[5:]:
                    acumulado += int(valor[i])                               
                sumatorio.append(str(acumulado))
            nueva_entrada.append(sumatorio)
            #print(nueva_entrada)

        fila_destino = diccionario['fila_destino']  # Fila donde empezar a pegar los datos
        
        if diccionario['folder']=="filter7":
            datos_transpuestos = list(map(list, zip(*nueva_entrada)))
        else:
            datos_transpuestos = nueva_entrada

        #print(datos_transpuestos)
        # Pegar los datos de la lista nueva_entrada en la hoja de Excel en la posición deseada
        for i, fila in enumerate(datos_transpuestos):
            for j, valor in enumerate(fila):            
                if diccionario['folder']=="filter6":
                    print('pasa por aquiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii filtro 6')

                    if j==1:
                        print(i,j,k,fila_destino+49+i,columna_destino_index + k ,valor)
                        celda = hoja_salida.cell(row=fila_destino+49+i, column=columna_destino_index + k -7 , value=valor)
                        celda.value = valor
                        celda.data_type = "n"  
                    else:            
                        celda = hoja_salida.cell(row=fila_destino+i, column=columna_destino_index + j + k-7  , value=valor)
                        celda.value = valor
                        celda.data_type = "n" 

                elif diccionario['folder']=="filter7":                 
                    i = i=12 if i==5 else i                   
                    print(i,j,k,fila_destino+j+i*7,columna_destino_index + k -14 ,valor)
                    celda = hoja_salida.cell(row=fila_destino+j+i*7, column=columna_destino_index+k-14, value=valor)
                    celda.value = valor
                    celda.data_type = "n"
                else:
                    print(i,j,k,fila_destino+i,columna_destino_index + j + k ,valor)
                                 
                    celda = hoja_salida.cell(row=fila_destino+i, column=columna_destino_index + j + k , value=valor)
                    celda.value = valor
                    celda.data_type = "n"

        #Guardar el libro de Excel con los datos pegados
        libro_salida.save(archivo_salida)
        k=k+1

def diccionarios():
    carpeta2 = "Entregable"
    archivo2 = "Indicadores de calidad.xlsx"
    archivo_salida = os.path.join(carpeta2, archivo2)
    hoja_nombre = "LIVE_diario"
    fila=2
    #columna_destino_index= 1283
    columna_destino_index=verify_column_empty(archivo_salida,hoja_nombre,fila)

    diccionarios = [
            {"name":"Unique Users (Suscribers).csv", "fila_destino": 3,"folder":"filter1"},
            {"name":"Effective Playtime.csv", "fila_destino" : 19,"folder":"filter1"},
            {"name":"Plays.csv", "fila_destino" : 43,"folder":"filter1"},
            {"name":"Buffer Ratio.csv","fila_destino" : 233,"folder":"filter1"},
            {"name":"In-Stream Error Crash (#).csv", "fila_destino": 257,"folder":"filter1"},
            {"name":"Happiness Score.csv" ,"fila_destino": 313,"folder":"filter1"},
            {"name":"Avg. Bitrate (Mbps).csv", "fila_destino" : 305,"folder":"filter1"},
            {"name":"Plays.csv", "fila_destino" : 51,"folder":"filter2"},
            {"name":"Buffer Ratio.csv","fila_destino" : 241,"folder":"filter2"},
            {"name":"In-Stream Error Crash (#).csv", "fila_destino": 265,"folder":"filter2"},
            {"name":"Plays.csv", "fila_destino" : 50,"folder":"filter3"},
            {"name":"Buffer Ratio.csv","fila_destino" : 240,"folder":"filter3"},
            {"name":"In-Stream Error Crash (#).csv", "fila_destino": 264,"folder":"filter3"},
            {"name":"Unique Users (Suscribers).csv", "fila_destino": 2,"folder":"filter4"},
            {"name":"Effective Playtime.csv", "fila_destino" : 18,"folder":"filter4"},
            {"name":"Plays.csv", "fila_destino" : 42,"folder":"filter4"},
            {"name":"Buffer Ratio.csv","fila_destino" : 232,"folder":"filter4"},
            {"name":"In-Stream Error Crash (#).csv", "fila_destino": 256,"folder":"filter4"},
            {"name":"Happiness Score.csv" ,"fila_destino": 312,"folder":"filter4"},
            {"name":"Avg. Bitrate (Mbps).csv", "fila_destino" : 304,"folder":"filter4"},          
           ]

    for diccionario in diccionarios:
        archivo_entrada = os.path.join(diccionario['folder'], diccionario['name'])
        rw_transponer(archivo_entrada,archivo_salida,columna_destino_index,diccionario['fila_destino'],diccionario['folder'])


def acumulado_seis(datos_entrada):
    nueva_entrada = datos_entrada[:5]  # Copiar los primeros cinco elementos de la lista original            
    acumulado = 0
    sumatorio=[]                         
    i=0
    for i in range(6):
        acumulado=0
        for valor in datos_entrada[5:]:
            acumulado += int(valor[i])                               
        sumatorio.append(str(acumulado))
    nueva_entrada.append(sumatorio)
    return nueva_entrada

def live_mensual():
    # Cargar el libro de Excel existente
    carpeta2 = "Entregable"
    archivo2 = "Indicadores de calidad.xlsx"
    archivo_salida = os.path.join(carpeta2, archivo2)

    #archivo_entrada = os.path.join("filter5", "Plays vs Join Time.csv")
    hoja_nombre = "LIVE_mensual"
    fila_inicio = 4  # Fila desde la cual iniciar la transposición o lectura
    columna_inicio = 3  # Columna desde la cual iniciar la transposición o lectura
    #columna_destino_index = 1282
    columna_destino_index=verify_column_empty(archivo_salida,hoja_nombre,fila=2)

    try:
        libro_salida = load_workbook(archivo_salida)
        nombres_hojas = libro_salida.sheetnames 
        hoja_indice = 2
        hoja_salida = libro_salida[nombres_hojas[hoja_indice]]
       

    except FileNotFoundError:
        libro_salida = Workbook()
        hoja_salida = libro_salida.active
        hoja_salida.title = hoja_nombre
    
    diccionarios = [
        {"name":"Performance Mensual Dispositivos_1.csv", "fila_destino": 3,"folder":"files_mensual"},
        {"name":"Performance Mensual Dispositivos_2.csv", "fila_destino" : 3,"folder":"files_mensual"},
        {"name":"Performance Mensual Datos - WIFI.csv", "fila_destino" : 3,"folder":"files_mensual"},
        {"name":"Performance Mensual Cabeceras.csv", "fila_destino" : 3,"folder":"files_mensual"}, 
    ]
    for diccionario in diccionarios:
        fila_destino = diccionario['fila_destino']  # Fila donde empezar a pegar los datos

        archivo_entrada = os.path.join(diccionario['folder'], diccionario['name'])
        #datos_entrada = liv3(archivo_entrada,fila_inicio,columna_inicio)
        if (diccionario['name']=='Performance Mensual Dispositivos_1.csv'):
            #print(datos_entrada)
            datos_entrada = lector_valores(archivo_entrada,fila_inicio,columna_inicio)

            i=0 #sirve para Dispositivos 1
            array=[3,19,43,233,257] #indices dispositivos 1

            j=0
            for valor in datos_entrada:      
                print(array[j]+i,columna_destino_index,valor)                   
                celda = hoja_salida.cell(row=array[j]+i, column=columna_destino_index, value=valor)
                celda.value = valor
                celda.data_type = "n"    
    
                if i == 6:             
                    i=0
                    j=j+1
                else:
                    i=i+1
        
        elif (diccionario['name']=='Performance Mensual Dispositivos_2.csv'):
            inicio_excluir = 22  # Índice de la primera fila a excluir
            fin_excluir = 37  # Índice de la última fila a excluir
            #print(datos_entrada)
            datos_entrada = excluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_excluir,fin_excluir)

            i=0 #sirve para Dispositivos 1
        
            array = [241,265,51] #indices dispositivos 2
            j=0
            for valor in datos_entrada:      
                print(array[j]+i,columna_destino_index,valor)                   
                celda = hoja_salida.cell(row=array[j]+i, column=columna_destino_index, value=valor)
                celda.value = valor
                celda.data_type = "n"    
    
                if i == 6:             
                    i=0
                    j=j+1
                else:
                    i=i+1
        
            #esto es dispositivos2
            inicio_incluir = 25  # Índice de la primera fila a excluir
            fin_incluir = 36  # Índice de la última fila a excluir
            datos_entrada = incluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_incluir,fin_incluir)

            datos_entrada=acumulado_seis(datos_entrada)
            print(datos_entrada)
            datos_entrada = list(map(list, zip(*datos_entrada)))
            print(datos_entrada)

            fila_destino=121 #esto es para dispositivos2 
            #columna_destino_index=1276
            for i, fila in enumerate(datos_entrada):
                print(i)
                for j, valor in enumerate(fila):
                    #print(valor)
                    if(i==5): #para saltar time o play 7 elementos
                        print(fila_destino+j+i*7+7*7,columna_destino_index,valor)   
                        celda = hoja_salida.cell(row=fila_destino+j+i*7+7*7, column=columna_destino_index, value=valor)
                        celda.value = valor
                        celda.data_type = "n" 
                    else:
                        print(fila_destino+j+i*7,columna_destino_index,valor)   
                        print('------------------------------------------------------------')                
                        celda = hoja_salida.cell(row=fila_destino+j+i*7, column=columna_destino_index, value=valor)
                        celda.value = valor
                        celda.data_type = "n" 
            libro_salida.save(archivo_salida)
        
        elif (diccionario['name']=='Performance Mensual Datos - WIFI.csv'):
            inicio_excluir=18
            fin_excluir=34
            array = []
            datos_entrada = excluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_excluir,fin_excluir)
            print(datos_entrada)
            
            array=[50,58,240,248,264,272] #indices wifi

            j=0
            for valor in datos_entrada:      
                print(array[j],columna_destino_index,valor)                   
                celda = hoja_salida.cell(row=array[j], column=columna_destino_index, value=valor)
                celda.value = valor
                celda.data_type = "n"
                j=j+1

            inicio_incluir=21
            fin_incluir=33
            
            datos_entrada = incluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_incluir,fin_incluir)
            print(datos_entrada)
            
            nueva_entrada = datos_entrada[:5]
            sumatorio=[]     
            for i in range(2):
                acumulado=0
                for valor in datos_entrada[5:]:
                    acumulado += int(valor[i])                               
                sumatorio.append(str(acumulado))
            nueva_entrada.append(sumatorio)
            print(nueva_entrada)
            
            array=[114,163]

            for i, fila in enumerate(nueva_entrada):
                for j, valor in enumerate(fila):
                    print(i,j,array[j]+i,columna_destino_index,valor)                   
                    celda = hoja_salida.cell(row=array[j]+i, column=columna_destino_index, value=valor)
                    celda.value = valor
                    celda.data_type = "n"
            libro_salida.save(archivo_salida)


        elif (diccionario['name']=='Performance Mensual Cabeceras.csv'):
            inicio_excluir = 35
            fin_excluir = 51
            datos_entrada = excluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_excluir,fin_excluir)
            #print(datos_entrada)
            #i=0 #sirve para Dispositivos 1
            array=[2,18,42,232,256,304,312] #indices dispositivos 1

            j=0
            for valor in datos_entrada:      
                #print(valor)
                print(array[j],columna_destino_index,valor)                   
                celda = hoja_salida.cell(row=array[j], column=columna_destino_index, value=valor)
                celda.value = valor
                celda.data_type = "n"
                j=j+1

            inicio_incluir = 38
            fin_incluir = 51
            datos_entrada = incluir_mayor_60s(archivo_entrada,fila_inicio,columna_inicio,inicio_incluir,fin_incluir)
            
            nueva_entrada = datos_entrada[:5]  # Copiar los primeros cinco elementos de la lista original            
            acumulado = 0
            sumatorio=[]                         
            i=0
            for i in range(1):
                acumulado=0
                for valor in datos_entrada[5:]:
                    acumulado += int(valor[i])                               
                sumatorio.append(str(acumulado))
            nueva_entrada.append(sumatorio)

            
            print(nueva_entrada)
            for i, fila in enumerate(nueva_entrada):
                #print(i)
                fila_destino=107
                for j, valor in enumerate(fila):
                    print(i,j,fila_destino+i,columna_destino_index,valor)
                    celda = hoja_salida.cell(row=fila_destino+i, column=columna_destino_index, value=valor)
                    celda.value = valor
                    celda.data_type = "n" 
         
       
            libro_salida.save(archivo_salida)
            

            # datos_entrada=acumulado_seis(datos_entrada)
            # print(datos_entrada)
            # datos_entrada = list(map(list, zip(*datos_entrada)))
            # print(datos_entrada)
            
            # for i, fila in enumerate(datos_entrada):
            #     print(i)
            #     for j, valor in enumerate(fila):
            #         print(fila_destino+i,columna_destino_index,valor)   
            
            # #incluir
                


def verify_column_empty(archivo_salida,nombre_hoja,fila):
    # Cargar el libro de Excel existente
    #carpeta2 = "Entregable"
    #archivo2 = "20230702- Indicadores de calidad.xlsx"
   
    #archivo_salida = os.path.join(carpeta2, archivo2)
    #nombre_hoja = 'LIVE_diario'

    # # Cargar el archivo de Excel
    libro = load_workbook(archivo_salida)

    # # Obtener la hoja especificada
    hoja = libro[nombre_hoja]

    # Buscar la primera celda vacía en la fila
    columna_vacia = None
    for columna in range(1, hoja.max_column + 1):
        valor_celda = hoja.cell(row=fila, column=columna).value
        if valor_celda is None:
            columna_vacia = columna
            break

    if columna_vacia:
        print(f"La primera columna vacía en la fila {fila} es la columna {columna_vacia}.")
    else:
        print(f"No se encontraron columnas vacías en la fila {fila}.")
    return columna_vacia

def reply_formula():
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import FORMULAE
    carpeta2 = "Entregable"
    archivo2 = "Indicadores de calidad.xlsx"

    archivo_salida = os.path.join(carpeta2, archivo2)
    libro = load_workbook(archivo_salida)
    # Seleccionar la hoja de trabajo
    hoja_nombre = "LIVE_diario"
    fila=1 #281 #10
    #columna_destino_index= 1283

    try:
        libro_salida = load_workbook(archivo_salida)
        nombres_hojas = libro_salida.sheetnames 
        hoja_indice = 3
        hoja = libro_salida[nombres_hojas[hoja_indice]] 
    except FileNotFoundError:
        #print(e)
        libro_salida = Workbook()
        # hoja_salida = libro_salida.active
        # hoja_salida.title = hoja_nombre

    columna_destino_index=verify_column_empty(archivo_salida,hoja_nombre,fila)
    print(columna_destino_index)
    letra_columna_i = get_column_letter(columna_destino_index-7)
    letra_columna_f = get_column_letter(columna_destino_index-1)
    letra_columna= get_column_letter(columna_destino_index)

    letra_columna_ia= get_column_letter(columna_destino_index-8)
    letra_columna_fd= get_column_letter(columna_destino_index+6)
    print('#############################vale 0###########################')    
 
    # #Esto sirve para arrastrar las formulas con la condicion que falta 1

    # # rangos = [
    # #     {"rango_origen":  hoja['AWA10:AWA17'], "rango_destino":hoja['AWB10:AWH17']},   #"rango_origen": hoja[letra_columna_ia+ str(10):letra_columna_ia+ str(17)],"rango_destino": hoja[letra_columna_i+ str(10):letra_columna_f+ str(17)]
    # #     {"rango_origen" : hoja['AWA288:AWA303'],"rango_destino" : hoja['AWB288:AWH303']},
    # #     {"rango_origen" : hoja['AWA113:AWA113'],"rango_destino" : hoja['AWB113:AWH113'] },
    # #     { "rango_origen" : hoja['AWA66:AWA106'],"rango_destino": hoja['AWB66:AWH106']},
    # #     {"rango_origen" : hoja['AWA26:AWA41'] ,"rango_destino": hoja['AWB26:AWH41']},
    # #     {"rango_origen" : hoja['AWA120:AWA120'] ,"rango_destino": hoja['AWB120:AWH120']},
    # #     {"rango_origen" : hoja['AWA127:AWA127'] ,"rango_destino": hoja['AWB127:AWH127']},
    # #     {"rango_origen" : hoja['AWA134:AWA134'] ,"rango_destino": hoja['AWB134:AWH134']},        
    # #     {"rango_origen" : hoja['AWA141:AWA141'] ,"rango_destino": hoja['AWB141:AWH141']},
    # #     {"rango_origen" : hoja['AWA148:AWA148'] ,"rango_destino": hoja['AWB148:AWH148']},
    # #     {"rango_origen" : hoja['AWA155:AWA155'] ,"rango_destino": hoja['AWB155:AWH155']},
    # #     {"rango_origen" : hoja['AWA162:AWA162'] ,"rango_destino": hoja['AWB162:AWH162']},
    # #     {"rango_origen" : hoja['AWA169:AWA169'] ,"rango_destino": hoja['AWB169:AWH169']},
    # #     {"rango_origen" : hoja['AWA176:AWA176'] ,"rango_destino": hoja['AWB176:AWH176']},
    # #     {"rango_origen" : hoja['AWA183:AWA183'] ,"rango_destino": hoja['AWB183:AWH183']},
    # #     {"rango_origen" : hoja['AWA190:AWA190'] ,"rango_destino": hoja['AWB190:AWH190']},
    # #     {"rango_origen" : hoja['AWA197:AWA197'] ,"rango_destino": hoja['AWB197:AWH197']},
    # #     {"rango_origen" : hoja['AWA204:AWA204'] ,"rango_destino": hoja['AWB204:AWH204']},    
    # #     {"rango_origen" : hoja['AWA280:AWA280'] ,"rango_destino": hoja['AWB280:AWH280']},    
    # #     ]

    rangos = [
        {"rango_origen": hoja[letra_columna_f+ str(10):letra_columna_f+ str(17)],"rango_destino": hoja[letra_columna+ str(10):letra_columna_fd+ str(17)]},       
        {"rango_origen" : hoja[letra_columna_f+ str(288):letra_columna_f+ str(303)],"rango_destino": hoja[letra_columna+ str(288):letra_columna_fd+ str(303)]},   
        {"rango_origen" : hoja[letra_columna_f+ str(113):letra_columna_f+ str(113)],"rango_destino": hoja[letra_columna+ str(113):letra_columna_fd+ str(113)] },  
        { "rango_origen" : hoja[letra_columna_f+ str(66):letra_columna_f+ str(106)],"rango_destino": hoja[letra_columna+ str(66):letra_columna_fd+ str(106)]},  
        {"rango_origen" : hoja[letra_columna_f+ str(26):letra_columna_f+ str(41)],"rango_destino": hoja[letra_columna+ str(26):letra_columna_fd+ str(41)]},  
        {"rango_origen" : hoja[letra_columna_f+ str(120):letra_columna_f+ str(120)],"rango_destino": hoja[letra_columna+ str(120):letra_columna_fd+ str(120)]},  
        {"rango_origen" : hoja[letra_columna_f+ str(127):letra_columna_f+ str(127)],"rango_destino": hoja[letra_columna+ str(127):letra_columna_fd+ str(127)]},  
        {"rango_origen" : hoja[letra_columna_f+ str(134):letra_columna_f+ str(134)],"rango_destino": hoja[letra_columna+ str(134):letra_columna_fd+ str(134)]},   
        {"rango_origen" : hoja[letra_columna_f+ str(141):letra_columna_f+ str(141)],"rango_destino": hoja[letra_columna+ str(141):letra_columna_fd+ str(141)]},  
        {"rango_origen" : hoja[letra_columna_f+ str(148):letra_columna_f+ str(148)],"rango_destino": hoja[letra_columna+ str(148):letra_columna_fd+ str(148)]},  
        {"rango_origen" : hoja[letra_columna_f+ str(155):letra_columna_f+ str(155)],"rango_destino": hoja[letra_columna+ str(155):letra_columna_fd+ str(155)]},
        {"rango_origen" : hoja[letra_columna_f+ str(162):letra_columna_f+ str(162)],"rango_destino": hoja[letra_columna+ str(162):letra_columna_fd+ str(162)]},
        {"rango_origen" : hoja[letra_columna_f+ str(169):letra_columna_f+ str(169)],"rango_destino": hoja[letra_columna+ str(169):letra_columna_fd+ str(169)]},
        {"rango_origen" : hoja[letra_columna_f+ str(176):letra_columna_f+ str(176)],"rango_destino": hoja[letra_columna+ str(176):letra_columna_fd+ str(176)]},
        {"rango_origen" : hoja[letra_columna_f+ str(183):letra_columna_f+ str(183)],"rango_destino": hoja[letra_columna+ str(183):letra_columna_fd+ str(183)]},
        {"rango_origen" : hoja[letra_columna_f+ str(190):letra_columna_f+ str(190)],"rango_destino": hoja[letra_columna+ str(190):letra_columna_fd+ str(190)]},
        {"rango_origen" : hoja[letra_columna_f+ str(197):letra_columna_f+ str(197)],"rango_destino": hoja[letra_columna+ str(197):letra_columna_fd+ str(197)]},
        {"rango_origen" : hoja[letra_columna_f+ str(204):letra_columna_f+ str(204)],"rango_destino": hoja[letra_columna+ str(204):letra_columna_fd+ str(204)]},   
        {"rango_origen" : hoja[letra_columna_f+ str(280):letra_columna_f+ str(280)],"rango_destino": hoja[letra_columna+ str(280):letra_columna_fd+ str(280)]}
    ]
    for r in rangos:
        #Recorrer el rango de celdas de origen y adaptar la fórmula a las celdas de destino
        for celda_origen, celda_destino in zip(r['rango_origen'], r['rango_destino']): 
            for destino in celda_destino:  
                formula = celda_origen[0].value            
                if formula and formula.startswith('='):              
                    nueva_formula = formula.replace(celda_origen[0].column_letter, get_column_letter(destino.column ))                                   
                    destino.value = nueva_formula
    # Recalcular las fórmulas en la hoja de trabajo
    hoja.calculate_dimension()
    # Guardar el archivo modificado
    libro_salida.save(archivo_salida)
    # print('###########################################################')


   
    #esto sirve para multiplica buffer radio y Avg Bitrate * 0.01 para mejorar la visualizacion
 
    rango_copia_h = [
        {"rango_origen": hoja[ letra_columna+ str(305): letra_columna_fd+str(311)]},
        {"rango_origen": hoja[ letra_columna+ str(232): letra_columna_fd+str(248)]}
        #{"rango_origen": hoja_salida["AWP305:AWV311"]},   
    ]

    for r in rango_copia_h:
        for celda_origen in r['rango_origen']:       
            for celda in celda_origen:
                celda = hoja.cell(row=celda.row, column=celda.column)
                celda.value = celda.value*0.01            
    libro_salida.save(archivo_salida)

    ################################ Sirve para copiar columnas de androidtv de celular a movil "####################################################################"
    
    #{"rango_origen": hoja[ letra_columna+ str(305): letra_columna_fd+str(311)]},

    rango_copia_h = [
        # {"rango_origen": hoja["AWB246:AWH246"],"rango_destino":hoja['AWB254:AWH254']},
        # {"rango_origen": hoja["AWB271:AWH271"],"rango_destino":hoja['AWB278:AWH278']} 
        {"rango_origen": hoja[letra_columna+ str(56): letra_columna_fd+str(56)],"rango_destino":hoja[letra_columna+ str(64): letra_columna_fd+ str(64)]},
        {"rango_origen": hoja[letra_columna+ str(246): letra_columna_fd+str(246)],"rango_destino":hoja[letra_columna+ str(254): letra_columna_fd+ str(254)]},
        {"rango_origen": hoja[letra_columna+ str(270): letra_columna_fd+str(270)],"rango_destino":hoja[letra_columna+ str(278): letra_columna_fd+ str(278)]},
    ]

    for r in rango_copia_h:
        i=0
        for celda_origen,celda_destino in zip(r['rango_origen'],r['rango_destino']):    
            print(celda_destino)
            print('destino')
            print(celda_origen)       
            for celda,destino in zip(celda_origen,celda_destino):      
                destino = hoja.cell(row=destino.row, column=destino.column)
                destino.value = celda.value            
    libro_salida.save(archivo_salida)

    print('#######################rrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr#########################################')

    # #Este codigo sirve para copia de una columna a las siguiente 7 
    # # rango_copia = [
    # #     {"rango_origen":   hoja["AWA59:AWA65"]}, hoja[letra_columna_i+ str(59):letra_columna_i+ str(65)]
    # #     {"rango_origen":   hoja["AWA156:AWA161"]}, hoja[letra_columna_i+ str(156):letra_columna_i+ str(161)]
    # #     {"rango_origen":   hoja["AWA170:AWA175"]}, hoja[letra_columna_i+ str(170):letra_columna_i+ str(175)]
    # #     {"rango_origen":   hoja["AWA177:AWA182"]}, hoja[letra_columna_i+ str(177):letra_columna_i+ str(182)]
    # #     {"rango_origen":   hoja["AWA184:AWA189"]}, hoja[letra_columna_i+ str(184):letra_columna_i+ str(189)]
    # #     {"rango_origen":   hoja["AWA191:AWA196"]}, hoja[letra_columna_i+ str(191):letra_columna_i+ str(196)]
    # #     {"rango_origen":   hoja["AWA198:AWA203"]}, hoja[letra_columna_i+ str(198):letra_columna_i+ str(203)]
    # #     {"rango_origen":   hoja["AWA273:AWA279"]}  hoja[letra_columna_i+ str(273):letra_columna_i+ str(279)]
    # # ]


    rango_copia = [
        {"rango_origen":   hoja[letra_columna_f+ str(59):letra_columna_f+ str(63)]},
        {"rango_origen":   hoja[letra_columna_f+ str(156):letra_columna_f+ str(161)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(170):letra_columna_f+ str(175)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(177):letra_columna_f+ str(182)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(184):letra_columna_f+ str(189)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(191):letra_columna_f+ str(196)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(198):letra_columna_f+ str(203)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(273):letra_columna_f+ str(277)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(279):letra_columna_f+ str(279)]}, 
        {"rango_origen":   hoja[letra_columna_f+ str(246):letra_columna_f+ str(246)]},  
        {"rango_origen":   hoja[letra_columna_f+ str(270):letra_columna_f+ str(270)]},  
        {"rango_origen":   hoja[letra_columna_f+ str(65):letra_columna_f+ str(65)]}, #  56 246
        {"rango_origen":   hoja[letra_columna_f+ str(56):letra_columna_f+ str(56)]}, #  56 246
    ]

    for r in rango_copia:
        for celda in zip(r['rango_origen']):       
            celda_origen=celda[0][0]
            columnas = 7  # Número de columnas a rellenar
            i=0
            for columna in range(2, columnas + 2):
                formula_origen = celda_origen.value               
                celda_destino = hoja.cell(row=celda_origen.row, column=celda_origen.column + columna - 1)
                celda_destino.value = formula_origen            
            i=i+1  
    # Guardar el archivo modificado
    libro_salida.save(archivo_salida)


    ############################ ultima formula =SI.ERROR(AWV257/AWU43,0)####################################
    rango_origen = hoja[letra_columna_f+ str(281):letra_columna_f+ str(287)]
    rango_destino= hoja[letra_columna+ str(281):letra_columna_fd+ str(287)]
 
    #Recorrer el rango de celdas de origen y adaptar la fórmula a las celdas de destino
    i=1    
    for celda_origen, celda_destino in zip(rango_origen, rango_destino):
        print(celda_origen)
        print(celda_destino)
        j=1
        for destino in celda_destino: 
            print('destino'+ str(destino))         
            formula = celda_origen[0].value
            print(formula)   
            #print(get_column_letter(columna_destino_index-(i)))        
            if formula and formula.startswith('='):              
                nueva_formula = formula.replace(celda_origen[0].column_letter, get_column_letter(destino.column ))                                   
                print(nueva_formula)
                letra_columna_f = get_column_letter(columna_destino_index+(j-2))
                print(letra_columna_f)
                nueva_formula2 = nueva_formula.replace('AWU', letra_columna_f) 
                print('ultima formula'+ nueva_formula2)                                  

                destino.value = nueva_formula2
            j=j+1
        print(i)
        i=i+1
            
    # Recalcular las fórmulas en la hoja de trabajo
    hoja.calculate_dimension()
    # Guardar el archivo modificado
    libro_salida.save(archivo_salida)

    # #############################Esto es para las cabeceras de fechas ##########################33
    dias = 7
    valor = hoja.cell(row=1, column=columna_destino_index-1).value
    print('valor'+str(valor))   
    # Rellenar las fechas en la columna A
   
    for i in range(dias):
        celda = hoja.cell(row=1, column=columna_destino_index+ i)
        print(i)
        celda.value = valor + timedelta(days=i+1)
        print(celda.value)

    libro_salida.save(archivo_salida)


    ############################despues de copiar eliminar de android tv ####################################################################
    ######### falta #########
    ##############################################################################################33

    ############################# limpiar las celdas copiadas ######################################

  

def reply_formula_mensual():
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import FORMULAE
    carpeta2 = "Entregable"
    archivo2 = "Indicadores de calidad.xlsx"

    archivo_salida = os.path.join(carpeta2, archivo2)
    libro = load_workbook(archivo_salida)
    # Seleccionar la hoja de trabajo
    hoja_nombre = "LIVE_mensual"
    fila=1 #es la primera fila desde donde se arrastra la copia
    columna_destino_index=verify_column_empty(archivo_salida,hoja_nombre,fila)
    print(columna_destino_index)
    letra_columna = get_column_letter(columna_destino_index)
    letra_columna_a = get_column_letter(columna_destino_index-1)

    try:
        libro_salida = load_workbook(archivo_salida)
        nombres_hojas = libro_salida.sheetnames 
        hoja_indice = 2
        hoja = libro_salida[nombres_hojas[hoja_indice]] 
    except FileNotFoundError:
        #print(e)
        libro_salida = Workbook()

    rango_copia = [
        
        {"rango_origen":   hoja[letra_columna_a+ str(280):letra_columna_a+ str(303)],"rango_destino": hoja[letra_columna+ str(280):letra_columna+ str(303)]},
        {"rango_origen":   hoja[letra_columna_a+ str(34):letra_columna_a+ str(34)],"rango_destino": hoja[letra_columna+ str(34):letra_columna+ str(34)]},
        {"rango_origen":   hoja[letra_columna_a+ str(90):letra_columna_a+ str(90)],"rango_destino": hoja[letra_columna+ str(90):letra_columna+ str(90)]},
        {"rango_origen":   hoja[letra_columna_a+ str(106):letra_columna_a+ str(106)],"rango_destino": hoja[letra_columna+ str(106):letra_columna+ str(106)]},
        {"rango_origen":   hoja[letra_columna_a+ str(113):letra_columna_a+ str(113)],"rango_destino": hoja[letra_columna+ str(113):letra_columna+ str(113)]},
        {"rango_origen":   hoja[letra_columna_a+ str(120):letra_columna_a+ str(120)],"rango_destino": hoja[letra_columna+ str(120):letra_columna+ str(120)]},
        {"rango_origen":   hoja[letra_columna_a+ str(127):letra_columna_a+ str(127)],"rango_destino": hoja[letra_columna+ str(127):letra_columna+ str(127)]},
        {"rango_origen":   hoja[letra_columna_a+ str(134):letra_columna_a+ str(134)],"rango_destino": hoja[letra_columna+ str(134):letra_columna+ str(134)]},
        {"rango_origen":   hoja[letra_columna_a+ str(141):letra_columna_a+ str(141)],"rango_destino": hoja[letra_columna+ str(141):letra_columna+ str(141)]},
        {"rango_origen":   hoja[letra_columna_a+ str(148):letra_columna_a+ str(148)],"rango_destino": hoja[letra_columna+ str(148):letra_columna+ str(148)]},
        {"rango_origen":   hoja[letra_columna_a+ str(162):letra_columna_a+ str(162)],"rango_destino": hoja[letra_columna+ str(162):letra_columna+ str(162)]},
        {"rango_origen":   hoja[letra_columna_a+ str(204):letra_columna_a+ str(204)],"rango_destino": hoja[letra_columna+ str(204):letra_columna+ str(204)]},
    ]

    for r in rango_copia:
        for celda_origen,celda_destino in zip(r['rango_origen'],r['rango_destino']):
            #Recorrer el rango de celdas de origen y adaptar la fórmula a las celdas de destino
            print(celda_origen)
            print(celda_destino)
            for destino in celda_destino:  
                formula = celda_origen[0].value  
                print(formula)          
                if formula and formula.startswith('='):              
                    nueva_formula = formula.replace(celda_origen[0].column_letter, get_column_letter(destino.column ))                                   
                    print(nueva_formula)
                    destino.value = nueva_formula
        # Recalcular las fórmulas en la hoja de trabajo
    hoja.calculate_dimension()
    # Guardar el archivo modificado
    libro_salida.save(archivo_salida)

    ###################################### Se multiplica por 0.01 ##################################################################
    # rango_copia_h = [
    #     {"rango_origen":   hoja[letra_columna_a+ str(232):letra_columna_a+ str(248)],"rango_destino": hoja[letra_columna+ str(232):letra_columna+ str(248)]},
    # ]
    rango_origen= hoja[letra_columna+ str(232):letra_columna+ str(248)]
    #for r in rango_copia_h:

    for celda_origen in rango_origen:       
        for celda in celda_origen:
            print('celda.row,celda.row')
            print(celda.row,celda.column)
            celda = hoja.cell(row=celda.row, column=celda.column)
            print('celda.value')
            print(celda)
            print(celda.value)
            celda.value = celda.value*0.01 

    libro_salida.save(archivo_salida)


    ############################# sirve para el autocompletado 0.0% ###########################################
    rango_copia = [
        {"rango_origen":   hoja[letra_columna_a+ str(59):letra_columna_a+ str(65)]},
        {"rango_origen":   hoja[letra_columna_a+ str(249):letra_columna_a+ str(255)]}, 
        {"rango_origen":   hoja[letra_columna_a+ str(273):letra_columna_a+ str(279)]}, 
    ]
    
    for r in rango_copia:
        for celda in zip(r['rango_origen']):       
            celda_origen=celda[0][0]
            columnas = 1  # Número de columnas a rellenar
            i=0
            for columna in range(2, columnas + 2):
                formula_origen = celda_origen.value               
                celda_destino = hoja.cell(row=celda_origen.row, column=celda_origen.column + columna - 1)
                celda_destino.value = formula_origen            
            i=i+1  
    # Guardar el archivo modificado
    libro_salida.save(archivo_salida)

    ####################################copia de buffer movil androidTv a wifi androidTV ##########################################################
    

    rango_copia_h = [ 
        {"rango_origen": hoja[letra_columna+ str(56): letra_columna+str(56)],"rango_destino":hoja[letra_columna+ str(64): letra_columna+ str(64)]},
        {"rango_origen": hoja[letra_columna+ str(246): letra_columna+str(246)],"rango_destino":hoja[letra_columna+ str(254): letra_columna+ str(254)]},
        {"rango_origen": hoja[letra_columna+ str(270): letra_columna+str(270)],"rango_destino":hoja[letra_columna+ str(278): letra_columna+ str(278)]},
    ]

    for r in rango_copia_h:
        i=0
        for celda_origen,celda_destino in zip(r['rango_origen'],r['rango_destino']):    
            print(celda_destino)
            print('destino')
            print(celda_origen)       
            for celda,destino in zip(celda_origen,celda_destino):      
                destino = hoja.cell(row=destino.row, column=destino.column)
                destino.value = celda.value            
    libro_salida.save(archivo_salida)

    ############################# limpiar las celdas copiadas ######################################

    rango_copia_h = [ 
        {"rango_origen": hoja[letra_columna_a+ str(56): letra_columna_a+str(56)],"rango_destino":hoja[letra_columna+ str(56): letra_columna+ str(56)]},
        {"rango_origen": hoja[letra_columna_a+ str(246): letra_columna_a+str(246)],"rango_destino":hoja[letra_columna+ str(246): letra_columna+ str(246)]},
        {"rango_origen": hoja[letra_columna_a+ str(270): letra_columna_a+str(270)],"rango_destino":hoja[letra_columna+ str(270): letra_columna+ str(270)]},
    ]

    for r in rango_copia_h:
        i=0
        for celda_origen,celda_destino in zip(r['rango_origen'],r['rango_destino']):    
            print(celda_destino)
            print('destino')
            print(celda_origen)       
            for celda,destino in zip(celda_origen,celda_destino):      
                destino = hoja.cell(row=destino.row, column=destino.column)
                destino.value = celda.value            
    libro_salida.save(archivo_salida)


    # #############################Esto es para las cabeceras de fechas ##########################33    
    valor = hoja.cell(row=1, column=columna_destino_index-1).value
    print('valor'+str(valor))   
    # Rellenar las fechas en la columna A


    # Agregar 1 mes a la fecha actual
    fecha_futura = valor + relativedelta(months=1)
    print(fecha_futura)   
    
    celda = hoja.cell(row=1, column=columna_destino_index) 
    celda.value = fecha_futura
    libro_salida.save(archivo_salida)

#configuration_youbora()
# filtro1_live()    
# filtro2_live()
# filtro3_live()
# filtro4_live()
#filtro5_live()
#filtro_mensual()
# diccionarios()
# plantilla()
#reply_formula()
#live_mensual()
#reply_formula_mensual()

#verify_column_empty()